from __future__ import annotations

import json
import statistics
from collections import Counter
from pathlib import Path

from .benchmark import run_benchmark
from .protocol import format_addr


def _latency_stats(values: list[float]) -> dict:
    if not values:
        return {"avg_ms": None, "min_ms": None, "max_ms": None, "p95_ms": None}
    sorted_values = sorted(values)
    p95_index = min(len(sorted_values) - 1, int(0.95 * (len(sorted_values) - 1)))
    return {
        "avg_ms": statistics.fmean(values),
        "min_ms": min(values),
        "max_ms": max(values),
        "p95_ms": sorted_values[p95_index],
    }


def aggregate_summaries(summaries: list[dict], latency_target_ms: float, loss_min: float, loss_max: float) -> dict:
    total_sent = sum(int(summary["total"]["sent"]) for summary in summaries)
    total_success = sum(int(summary["total"]["success"]) for summary in summaries)
    all_latencies = [
        float(round_record["latency_ms"])
        for summary in summaries
        for round_record in summary.get("rounds", [])
        if round_record.get("success") and round_record.get("latency_ms") is not None
    ]
    targets: dict[str, dict] = {}
    route_counter: Counter[str] = Counter()
    for summary in summaries:
        for target, item in summary.get("targets", {}).items():
            target_data = targets.setdefault(target, {"sent": 0, "success": 0, "lost": 0, "latencies": []})
            target_data["sent"] += int(item.get("sent", 0))
            target_data["success"] += int(item.get("success", 0))
            target_data["lost"] += int(item.get("lost", 0))
            route = item.get("route_path") or ""
            if route:
                route_counter[f"{target}|{route}"] += 1
        for round_record in summary.get("rounds", []):
            if round_record.get("success") and round_record.get("latency_ms") is not None:
                target = f"{int(round_record['target']):02X}"
                targets.setdefault(target, {"sent": 0, "success": 0, "lost": 0, "latencies": []})["latencies"].append(
                    float(round_record["latency_ms"])
                )

    target_summary = {}
    for target, item in sorted(targets.items()):
        sent = int(item["sent"])
        success = int(item["success"])
        target_summary[target] = {
            "sent": sent,
            "success": success,
            "lost": sent - success,
            "loss_rate": (sent - success) / sent if sent else None,
            "latency": _latency_stats(item.get("latencies", [])),
        }

    loss_rate = (total_sent - total_success) / total_sent if total_sent else None
    latency = _latency_stats(all_latencies)
    return {
        "schema": "pc_d3qn_cli.batch_benchmark_summary.v1",
        "runs": [
            {
                "index": index + 1,
                "log_dir": summary.get("config", {}).get("log_dir") or summary.get("log_dir"),
                "sent": summary["total"]["sent"],
                "success": summary["total"]["success"],
                "loss_rate": summary["total"]["loss_rate"],
                "avg_latency_ms": summary["total"]["latency"]["avg_ms"],
                "p95_latency_ms": summary["total"]["latency"]["p95_ms"],
                "route_failures": summary.get("d3qn_route_failures", 0),
            }
            for index, summary in enumerate(summaries)
        ],
        "total": {
            "sent": total_sent,
            "success": total_success,
            "lost": total_sent - total_success,
            "loss_rate": loss_rate,
            "latency": latency,
            "route_failures": sum(int(summary.get("d3qn_route_failures", 0)) for summary in summaries),
        },
        "targets": target_summary,
        "path_frequency": [
            {"target": key.split("|", 1)[0], "route": key.split("|", 1)[1], "count": count}
            for key, count in route_counter.most_common()
        ],
        "goal": {
            "avg_latency_target_ms": latency_target_ms,
            "loss_rate_min": loss_min,
            "loss_rate_max": loss_max,
            "latency_met": latency["avg_ms"] is not None and latency["avg_ms"] <= latency_target_ms,
            "loss_met": loss_rate is not None and loss_min <= loss_rate <= loss_max,
        },
    }


def _fmt_ms(value) -> str:
    return "n/a" if value is None else f"{float(value):.1f}ms"


def _fmt_rate(value) -> str:
    return "n/a" if value is None else f"{float(value):.2%}"


def write_batch_report(path: str | Path, summary: dict) -> None:
    goal = summary["goal"]
    total = summary["total"]
    lines = [
        "# D3QN_MPNN 10次硬件测试总汇报",
        "",
        "- 算法：`D3QN_MPNN`",
        "- 推理策略：`纯D3QN，无Dijkstra fallback，无规则兜底`",
        f"- 总发送：`{total['sent']}`，成功：`{total['success']}`，失败：`{total['lost']}`",
        f"- 整体丢包率：`{_fmt_rate(total['loss_rate'])}`，目标：`{_fmt_rate(goal['loss_rate_min'])}~{_fmt_rate(goal['loss_rate_max'])}`，达标：`{'是' if goal['loss_met'] else '否'}`",
        f"- 整体平均延时：`{_fmt_ms(total['latency']['avg_ms'])}`，目标：`{_fmt_ms(goal['avg_latency_target_ms'])}`，达标：`{'是' if goal['latency_met'] else '否'}`",
        f"- P95延时：`{_fmt_ms(total['latency']['p95_ms'])}`",
        f"- D3QN路由失败次数：`{total['route_failures']}`",
        "",
        "## 单次测试对比",
        "",
        "| 次数 | 日志目录 | 发送 | 成功 | 丢包率 | 平均延时 | P95 | 路由失败 |",
        "|---:|---|---:|---:|---:|---:|---:|---:|",
    ]
    for run in summary["runs"]:
        lines.append(
            f"| {run['index']} | `{run['log_dir']}` | {run['sent']} | {run['success']} | "
            f"{_fmt_rate(run['loss_rate'])} | {_fmt_ms(run['avg_latency_ms'])} | {_fmt_ms(run['p95_latency_ms'])} | {run['route_failures']} |"
        )
    lines.extend([
        "",
        "## 各节点聚合",
        "",
        "| 目标点 | 成功/发送 | 丢包率 | 平均延时 | P95 |",
        "|---|---:|---:|---:|---:|",
    ])
    for target, item in sorted(summary["targets"].items()):
        lines.append(
            f"| `{target}` | {item['success']}/{item['sent']} | {_fmt_rate(item['loss_rate'])} | "
            f"{_fmt_ms(item['latency']['avg_ms'])} | {_fmt_ms(item['latency']['p95_ms'])} |"
        )
    lines.extend([
        "",
        "## 路径选择频次",
        "",
        "| 目标点 | 路径 | 次数 |",
        "|---|---|---:|",
    ])
    for item in summary["path_frequency"][:50]:
        lines.append(f"| `{item['target']}` | `{item['route']}` | {item['count']} |")
    Path(path).write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_batch_excel(path: str | Path, summary: dict) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        Path(path).with_suffix(".csv").write_text("openpyxl unavailable\n", encoding="utf-8")
        return
    workbook = Workbook()
    ws = workbook.active
    ws.title = "10次测试汇总"
    ws.append(["次数", "日志目录", "发送", "成功", "丢包率", "平均延时ms", "P95ms", "路由失败"])
    for run in summary["runs"]:
        ws.append([
            run["index"],
            run["log_dir"],
            run["sent"],
            run["success"],
            run["loss_rate"],
            run["avg_latency_ms"],
            run["p95_latency_ms"],
            run["route_failures"],
        ])
    ws2 = workbook.create_sheet("各节点聚合")
    ws2.append(["目标点", "发送", "成功", "失败", "丢包率", "平均延时ms", "P95ms"])
    for target, item in sorted(summary["targets"].items()):
        ws2.append([target, item["sent"], item["success"], item["lost"], item["loss_rate"], item["latency"]["avg_ms"], item["latency"]["p95_ms"]])
    ws3 = workbook.create_sheet("路径频次")
    ws3.append(["目标点", "路径", "次数"])
    for item in summary["path_frequency"]:
        ws3.append([item["target"], item["route"], item["count"]])
    workbook.save(path)


def run_batch_benchmark(
    *,
    repeats: int,
    output_dir: str | Path,
    latency_target_ms: float,
    loss_min: float,
    loss_max: float,
    **benchmark_kwargs,
) -> dict:
    summaries = []
    for index in range(1, repeats + 1):
        print(f"batch benchmark {index}/{repeats}")
        summaries.append(run_benchmark(**benchmark_kwargs))
    summary = aggregate_summaries(summaries, latency_target_ms, loss_min, loss_max)
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    (output / "10次测试汇总.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    write_batch_report(output / "10次测试总汇报.md", summary)
    write_batch_excel(output / "10次测试汇总.xlsx", summary)
    return summary
